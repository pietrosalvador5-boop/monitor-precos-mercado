import asyncio
import csv
import json
import os
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from urllib.parse import quote, quote_plus

import pandas as pd
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

from normalizacao import extrair_precos, normalizar_texto, score_match

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
SAIDAS_DIR = BASE_DIR / "saidas"
SAIDAS_DIR.mkdir(exist_ok=True)

ARQUIVO_ITENS = DATA_DIR / "itens_monitorados.csv"
ARQUIVO_HISTORICO = SAIDAS_DIR / "historico_precos_long.csv"
ARQUIVO_JSON_ERROS = SAIDAS_DIR / "erros_ultima_coleta.json"

LOJAS = {
    "zaffari": {
        "nome": "Zaffari",
        "nome_planilha": "Zaffari",
        "url_inicial": "https://www.zaffari.com.br/",
        "busca_url": "https://www.zaffari.com.br/{query}?_q={query}&map=ft",
        "usar_busca_visual": False,
    },
    "pao_de_acucar": {
        "nome": "Pão de Açúcar",
        "nome_planilha": "Pao_de_Acucar",
        "url_inicial": "https://www.paodeacucar.com/",
        "busca_url": "https://www.paodeacucar.com/busca?terms={query}",
        "usar_busca_visual": False,
    },
}

CARD_SELECTORS = [
    "article",
    "li",
    "[data-testid*='product']",
    "[class*='product']",
    "[class*='Product']",
    "[class*='card']",
    "[class*='Card']",
]

@dataclass
class Resultado:
    data_coleta: str
    loja: str
    item: str
    termo_busca: str
    unidade_padrao: str
    produto_encontrado: str = ""
    preco: Optional[float] = None
    url: str = ""
    score: float = 0.0
    status: str = ""
    observacao: str = ""
    qtd_precos_no_card: int = 0


def carregar_itens():
    with open(ARQUIVO_ITENS, encoding="utf-8") as f:
        return list(csv.DictReader(f))


async def aceitar_cookies(page):
    textos = ["Aceitar", "Aceito", "Entendi", "Permitir", "Continuar", "Concordo"]
    for texto in textos:
        try:
            btn = page.get_by_role("button", name=texto)
            if await btn.count() > 0:
                await btn.first.click(timeout=2500)
                return
        except Exception:
            pass


async def abrir_busca(page, loja_cfg, termo):
    if loja_cfg.get("busca_url"):
        url = loja_cfg["busca_url"].format(query=quote(termo))
        await page.goto(url, wait_until="domcontentloaded", timeout=45000)
    else:
        await page.goto(loja_cfg["url_inicial"], wait_until="domcontentloaded", timeout=45000)
        await aceitar_cookies(page)
        # Busca visual genérica: tenta achar um campo de pesquisa.
        possiveis = [
            "input[type='search']",
            "input[placeholder*='Buscar']",
            "input[placeholder*='buscar']",
            "input[placeholder*='Pesquisar']",
            "input[placeholder*='Pesquise']",
            "input[name*='search']",
            "input[name*='busca']",
            "input",
        ]
        campo = None
        for sel in possiveis:
            try:
                loc = page.locator(sel)
                if await loc.count() > 0:
                    campo = loc.first
                    break
            except Exception:
                continue
        if campo is None:
            raise RuntimeError("Campo de busca não encontrado no site.")
        await campo.wait_for(state="visible", timeout=30000)
        await campo.click(timeout=30000)
        await campo.fill(termo, timeout=30000)
        await campo.press("Enter", timeout=30000)
    try:
        await page.wait_for_selector("text=R$", timeout=15000)
    except Exception:
        pass
    await page.wait_for_timeout(3000)


async def extrair_cards(page, termo, max_cards=40):
    html = await page.content()
    soup = BeautifulSoup(html, "lxml")
    candidatos = []

    for selector in CARD_SELECTORS:
        for el in soup.select(selector):
            texto = " ".join(el.get_text(" ", strip=True).split())
            if "R$" not in texto or len(texto) < 8 or len(texto) > 1200:
                continue
            precos = extrair_precos(texto)
            if not precos:
                continue
            link = ""
            a = el.select_one("a[href]")
            if a:
                link = a.get("href", "")
            candidatos.append({
                "texto": texto,
                "preco": min(precos),
                "qtd_precos": len(precos),
                "score": score_match(termo, texto),
                "link": link,
            })

    # Remove duplicatas por texto normalizado + preço
    unicos = {}
    for c in candidatos:
        chave = (normalizar_texto(c["texto"])[:120], c["preco"])
        if chave not in unicos or c["score"] > unicos[chave]["score"]:
            unicos[chave] = c

    ordenados = sorted(unicos.values(), key=lambda x: (-x["score"], x["preco"]))
    return ordenados[:max_cards]


async def coletar_item(context, loja_id, loja_cfg, item_row):
    item = item_row["item"]
    termo = item_row.get("termo_busca") or item
    unidade = item_row.get("unidade_padrao", "")
    data_coleta = date.today().isoformat()

    if item_row.get("coletar", "sim").strip().lower() != "sim":
        return Resultado(
            data_coleta=data_coleta,
            loja=loja_cfg["nome"],
            item=item,
            termo_busca=termo,
            unidade_padrao=unidade,
            status="nao_coletado",
            observacao=item_row.get("observacao", "Item marcado para não coletar automaticamente"),
        )

    page = await context.new_page()
    try:
        await abrir_busca(page, loja_cfg, termo)
        await aceitar_cookies(page)
        cards = await extrair_cards(page, termo)
        url_atual = page.url
        if not cards:
            return Resultado(data_coleta, loja_cfg["nome"], item, termo, unidade, url=url_atual, status="nao_encontrado")
        melhor = cards[0]
        # Se o card tem mais de um preço (ex: combo, "de X por Y", produto +
        # acompanhamento), não dá para ter certeza de qual preço é o do item
        # monitorado. Em vez de assumir o menor valor, marcamos para revisão
        # manual e mantemos o menor preço apenas como sugestão/observação.
        ambiguo = melhor["qtd_precos"] > 1
        if melhor["score"] >= 0.35 and not ambiguo:
            status = "ok"
        elif melhor["score"] >= 0.35 and ambiguo:
            status = "revisar_match"
        else:
            status = "revisar_match"
        observ = ""
        if ambiguo:
            observ = f"Card continha {melhor['qtd_precos']} valores em R$; preço escolhido é o menor encontrado, confirme manualmente."
        return Resultado(
            data_coleta=data_coleta,
            loja=loja_cfg["nome"],
            item=item,
            termo_busca=termo,
            unidade_padrao=unidade,
            produto_encontrado=melhor["texto"][:500],
            preco=melhor["preco"],
            url=url_atual if not melhor["link"] else melhor["link"],
            score=round(melhor["score"], 3),
            status=status,
            observacao=observ,
            qtd_precos_no_card=melhor["qtd_precos"],
        )
    except PlaywrightTimeoutError as e:
        return Resultado(data_coleta, loja_cfg["nome"], item, termo, unidade, status="erro_timeout", observacao=str(e)[:250])
    except Exception as e:
        return Resultado(data_coleta, loja_cfg["nome"], item, termo, unidade, status="erro", observacao=str(e)[:250])
    finally:
        await page.close()


def salvar_historico(resultados):
    df_novo = pd.DataFrame([r.__dict__ for r in resultados])
    if ARQUIVO_HISTORICO.exists():
        df_antigo = pd.read_csv(ARQUIVO_HISTORICO)
        df = pd.concat([df_antigo, df_novo], ignore_index=True)
        df = df.drop_duplicates(subset=["data_coleta", "loja", "item"], keep="last")
    else:
        df = df_novo
    df.to_csv(ARQUIVO_HISTORICO, index=False, encoding="utf-8-sig")
    return df


def _localizar_linha_cabecalho(ws, max_linhas_busca=10):
    """Procura a linha que contém 'Data da coleta' na coluna A.
    A planilha-base tem título e instruções nas linhas 1-3, cabeçalho real na linha 4."""
    for r in range(1, max_linhas_busca + 1):
        valor = ws.cell(row=r, column=1).value
        if isinstance(valor, str) and valor.strip().lower() in ("data da coleta", "data"):
            return r
    return None


def atualizar_excel(caminho_excel: str, df_historico: pd.DataFrame):
    """Atualiza a planilha-base em formato largo: linha=data (semana pré-criada), colunas=itens.

    Respeita a estrutura real do arquivo: título na linha 1, instruções na linha 2,
    cabeçalho ('Data da coleta', 'Semana', <itens...>) na linha 4, e linhas de semana
    já pré-criadas com datas futuras. Nunca cria linha nova nem sobrescreve o título;
    se a data de hoje não bater com nenhuma semana pré-criada, registra aviso e pula.
    """
    from openpyxl import load_workbook

    caminho = Path(caminho_excel)
    if not caminho.exists():
        print(f"Planilha não encontrada: {caminho}")
        return

    wb = load_workbook(caminho)
    data_hoje = date.today().isoformat()
    avisos = []

    for loja_id, loja_cfg in LOJAS.items():
        nome_sheet = loja_cfg["nome_planilha"]
        if nome_sheet not in wb.sheetnames:
            continue
        ws = wb[nome_sheet]

        linha_cabecalho = _localizar_linha_cabecalho(ws)
        if linha_cabecalho is None:
            avisos.append(f"{nome_sheet}: não encontrei a linha de cabeçalho ('Data da coleta'); pulei esta aba.")
            continue

        headers = [ws.cell(row=linha_cabecalho, column=c).value for c in range(1, ws.max_column + 1)]

        # Acrescenta apenas itens que ainda não existem como coluna (não deveria
        # ser necessário se a planilha já foi gerada com a lista completa, mas
        # evita perder dados de itens novos adicionados ao CSV depois).
        itens = list(pd.read_csv(ARQUIVO_ITENS)["item"])
        for item in itens:
            if item not in headers:
                nova_col = ws.max_column + 1
                ws.cell(row=linha_cabecalho, column=nova_col).value = item
                headers.append(item)

        # Procura a linha cuja data (coluna A) corresponde a hoje, dentro das
        # semanas pré-criadas. Não cria linha nova — a planilha já vem com 52
        # semanas prontas para o ano.
        linha_data = None
        for r in range(linha_cabecalho + 1, ws.max_row + 1):
            valor = ws.cell(row=r, column=1).value
            if valor is None:
                continue
            valor_iso = valor.date().isoformat() if hasattr(valor, "date") else str(valor)[:10]
            if valor_iso == data_hoje:
                linha_data = r
                break

        if linha_data is None:
            avisos.append(
                f"{nome_sheet}: não há uma linha de 'Semana' pré-criada para a data {data_hoje}. "
                f"Os preços desta coleta NÃO foram gravados na aba {nome_sheet} (ficaram só em "
                f"saidas/historico_precos_long.csv). Adicione mais linhas de semana na planilha-base."
            )
            continue

        recortes = df_historico[(df_historico["data_coleta"] == data_hoje) & (df_historico["loja"] == loja_cfg["nome"])]
        mapa = dict(zip(recortes["item"], recortes["preco"]))
        for item, preco in mapa.items():
            if item in headers and pd.notna(preco):
                col = headers.index(item) + 1
                ws.cell(row=linha_data, column=col).value = float(preco)

    if avisos:
        print("AVISOS ao atualizar a planilha:")
        for a in avisos:
            print(" -", a)

    # Aba detalhada para auditoria
    if "Coletas_Detalhadas" in wb.sheetnames:
        del wb["Coletas_Detalhadas"]
    ws = wb.create_sheet("Coletas_Detalhadas")
    cols = list(df_historico.columns)
    ws.append(cols)
    for _, row in df_historico.iterrows():
        ws.append([row.get(c, "") for c in cols])
    wb.save(caminho)
    print(f"Planilha atualizada: {caminho}")


async def main():
    itens = carregar_itens()
    resultados = []
    erros = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=os.getenv("HEADLESS", "true").lower() != "false")
        context = await browser.new_context(locale="pt-BR", timezone_id="America/Sao_Paulo")
        for loja_id, loja_cfg in LOJAS.items():
            for i, item_row in enumerate(itens, 1):
                r = await coletar_item(context, loja_id, loja_cfg, item_row)
                resultados.append(r)
                print(f"[{loja_cfg['nome']}] {i}/{len(itens)} - {r.item}: {r.status} {r.preco or ''}")
                if r.status.startswith("erro"):
                    erros.append(r.__dict__)
                await asyncio.sleep(float(os.getenv("SLEEP_SECONDS", "2")))
        await browser.close()

    df = salvar_historico(resultados)
    with open(ARQUIVO_JSON_ERROS, "w", encoding="utf-8") as f:
        json.dump(erros, f, ensure_ascii=False, indent=2)

    caminho_excel = os.getenv("CAMINHO_EXCEL")
    if caminho_excel:
        atualizar_excel(caminho_excel, df)

if __name__ == "__main__":
    asyncio.run(main())
